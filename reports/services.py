"""PDF (reportlab) ve Excel (openpyxl) rapor üretim servisleri."""
import io

from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# ---------------------------------------------------------------------------
# Türkçe karakter desteği (ğ, ş, ı, İ, ö, ü, ç):
# ReportLab'ın yerleşik (core) fontları -- Helvetica, Times-Roman vb. -- WinAnsi
# kodlamasını kullanır ve bu kodlamada ğ, ş, ı, İ karakterleri YOKTUR. Bu yüzden
# PDF'lerde (özellikle öğrenci karnesinde) bu harfler eksik/bozuk çıkıyordu.
# Projede zaten static/fonts/ altında bulunan (ama hiç kullanılmayan) Unicode
# destekli DejaVu Sans fontlarını burada ReportLab'a kaydedip tüm PDF
# metinlerinde ve tablolarda bu fontları kullanıyoruz.
# ---------------------------------------------------------------------------
FONT_REGULAR = "DejaVuSans"
FONT_BOLD = "DejaVuSans-Bold"

if FONT_REGULAR not in pdfmetrics.getRegisteredFontNames():
    _fonts_dir = settings.BASE_DIR / "static" / "fonts"
    pdfmetrics.registerFont(TTFont(FONT_REGULAR, str(_fonts_dir / "DejaVuSans.ttf")))
    pdfmetrics.registerFont(TTFont(FONT_BOLD, str(_fonts_dir / "DejaVuSans-Bold.ttf")))


def _tr_styles():
    """getSampleStyleSheet()'in Türkçe karakterleri doğru gösteren (DejaVu Sans
    fontlu) sürümü. Başlık/gövde stillerinin fontName'i burada değiştirilir."""
    styles = getSampleStyleSheet()
    styles["Title"].fontName = FONT_BOLD
    styles["Normal"].fontName = FONT_REGULAR
    styles["Heading3"].fontName = FONT_BOLD
    return styles


def _table_style(extra=None):
    base = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#065f46")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
        ("FONTNAME", (0, 1), (-1, -1), FONT_REGULAR),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    if extra:
        base.extend(extra)
    return TableStyle(base)


# Çok sütunlu raporlarda (ör. hedef ilerleme sütunları eklenince) hücrelerin
# sayfa dışına taşmaması için başlık/hücre metinlerini Paragraph içine alıp
# satır kaydırmalı (word-wrap) hale getiriyoruz.
_HEADER_CELL_STYLE = ParagraphStyle(
    "TableHeaderCell", fontName=FONT_BOLD, fontSize=7.5, leading=9, textColor=colors.white,
)
_BODY_CELL_STYLE = ParagraphStyle(
    "TableBodyCell", fontName=FONT_REGULAR, fontSize=7.5, leading=9,
)


def build_pdf_table(title, headers, rows, subtitle=None):
    buffer = io.BytesIO()
    left_margin = right_margin = 1 * cm
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4), topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=left_margin, rightMargin=right_margin,
    )
    styles = _tr_styles()
    elements = [Paragraph(title, styles["Title"])]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    elements.append(Spacer(1, 0.5 * cm))

    raw_rows = rows if rows else [["Kayıt bulunamadı"] + [""] * (len(headers) - 1)]
    header_row = [Paragraph(str(h), _HEADER_CELL_STYLE) for h in headers]
    body_rows = [
        [Paragraph("-" if v in (None, "") else str(v), _BODY_CELL_STYLE) for v in row]
        for row in raw_rows
    ]
    data = [header_row] + body_rows

    available_width = landscape(A4)[0] - left_margin - right_margin
    col_width = available_width / len(headers)
    table = Table(data, colWidths=[col_width] * len(headers), repeatRows=1)
    table.setStyle(_table_style())
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def build_excel_table(title, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] if title else "Rapor"

    ws.append(headers)
    header_fill = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for row in rows:
        ws.append(row)

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def lesson_rows_and_headers(lessons):
    headers = ["Öğrenci", "Tarih", "Devam", "Ham (Cüz)", "Tekrar Edilen Cüzler", "Kalite", "Not"]
    rows = []
    for lesson in lessons:
        rows.append([
            lesson.student.full_name,
            lesson.date.strftime("%d.%m.%Y"),
            lesson.get_attendance_display(),
            lesson.ham_juz_label or "-",
            ", ".join(lesson.revision_juz_labels) or "-",
            lesson.get_quality_display() if lesson.quality else "-",
            (lesson.notes or "")[:60],
        ])
    return headers, rows


def attendance_rows_and_headers(students):
    from lessons.models import LessonRecord
    headers = ["Öğrenci", "Toplam Ders", "Geldi", "Gelmedi", "İzinli", "Devam Yüzdesi (%)"]
    rows = []
    for s in students:
        qs = LessonRecord.objects.filter(student=s)
        total = qs.count()
        present = qs.filter(attendance=LessonRecord.Attendance.PRESENT).count()
        absent = qs.filter(attendance=LessonRecord.Attendance.ABSENT).count()
        excused = qs.filter(attendance=LessonRecord.Attendance.EXCUSED).count()
        pct = round((present / total) * 100, 1) if total else 0
        rows.append([s.full_name, total, present, absent, excused, pct])
    return headers, rows


def progress_rows_and_headers(students):
    from memorization.services import get_progress_summary, get_juz_progress_summary
    from predictions.services import calculate_target_progress
    headers = [
        "Öğrenci", "Tamamlanan Cüz", "Tekrar Bekleyen Cüz", "Çalışılmamış Cüz", "İlerleme (%)",
        "Hedef Tarih", "Hedefe Göre Beklenen (Sayfa)", "Hedef İlerleme (%)", "Gereken Haftalık Tempo (Sayfa)",
    ]
    rows = []
    for s in students:
        summary = get_progress_summary(s)
        juz_summary = get_juz_progress_summary(s)
        target = calculate_target_progress(s)
        if target:
            target_cols = [
                s.target_completion_date.strftime("%d.%m.%Y"),
                target["expected_pages_by_now"],
                f"%{target['target_progress_percent']}",
                target["required_weekly_pace"] if target["required_weekly_pace"] is not None else "-",
            ]
        else:
            target_cols = ["-", "-", "-", "-"]
        rows.append([
            s.full_name,
            f"{juz_summary['completed_juz']} / {juz_summary['total_juz']}",
            juz_summary["needs_revision_juz"],
            juz_summary["not_studied_juz"],
            summary["progress_percent"],
            *target_cols,
        ])
    return headers, rows


def prediction_rows_and_headers(students):
    headers = ["Öğrenci", "Kalan Sayfa", "Tahmini Kalan Gün", "Tahmini Bitiş Tarihi", "Güven Seviyesi", "Yöntem"]
    rows = []
    for s in students:
        pred = s.predictions.first()
        if not pred:
            rows.append([s.full_name, "-", "-", "-", "-", "-"])
            continue
        rows.append([
            s.full_name, pred.remaining_pages, pred.estimated_remaining_days or "-",
            pred.estimated_completion_date.strftime("%d.%m.%Y") if pred.estimated_completion_date else "-",
            pred.get_confidence_level_display(), pred.get_method_used_display(),
        ])
    return headers, rows


def build_student_report_card_pdf(student):
    """Tek bir öğrenci için 'karne' niteliğinde birleşik PDF rapor (ilerleme + devam + tahmin + son dersler)."""
    from django.utils import timezone
    from memorization.services import get_progress_summary, get_juz_progress_summary
    from lessons.services import get_attendance_summary
    from lessons.models import LessonRecord
    from predictions.services import calculate_target_progress

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = _tr_styles()
    elements = [
        Paragraph(f"Öğrenci Karnesi — {student.full_name}", styles["Title"]),
        Paragraph(
            f"Hafızlığa başlama: {student.start_date.strftime('%d.%m.%Y')} · "
            f"Öğretici: {student.teacher.get_full_name() or student.teacher.username} · "
            f"Rapor tarihi: {timezone.localdate().strftime('%d.%m.%Y')}",
            styles["Normal"],
        ),
        Spacer(1, 0.6 * cm),
    ]

    progress = get_progress_summary(student)
    juz_progress = get_juz_progress_summary(student)
    attendance = get_attendance_summary(student)
    prediction = student.predictions.first()

    summary_headers = ["Ölçüt", "Değer"]
    summary_rows = [
        ["İlerleme Yüzdesi", f"%{progress['progress_percent']}"],
        ["Tamamlanan Cüz", f"{juz_progress['completed_juz']} / {juz_progress['total_juz']}"],
        ["Tekrar Bekleyen Cüz", str(juz_progress["needs_revision_juz"])],
        ["Devam Yüzdesi", f"%{attendance['attendance_percent']}"],
        ["Toplam Devamsızlık", str(attendance["absent"])],
        ["Bu Ay Devamsızlık", str(attendance["monthly_absent"])],
        ["Ardışık Devamsızlık", str(attendance["consecutive_absent"])],
    ]
    if prediction and prediction.estimated_completion_date:
        summary_rows += [
            ["Tahmini Bitiş Tarihi", prediction.estimated_completion_date.strftime("%d.%m.%Y")],
            ["Tahmini Kalan Gün", str(prediction.estimated_remaining_days)],
            ["Tahmin Güven Seviyesi", prediction.get_confidence_level_display()],
        ]

    target_progress = calculate_target_progress(student)
    if target_progress:
        summary_rows += [
            ["Hedef Bitiş Tarihi", student.target_completion_date.strftime("%d.%m.%Y")],
            ["Hedefe Göre Bugün Olması Gereken", f"{target_progress['expected_pages_by_now']} sayfa (~{target_progress['expected_juz_by_now']}. Cüz)"],
            ["Hedefe Göre İlerleme", f"%{target_progress['target_progress_percent']}"],
            [
                "Hedeften Fark",
                (f"{target_progress['pages_ahead_behind_abs']} sayfa önde" if target_progress["pages_ahead_behind"] > 0
                 else f"{target_progress['pages_ahead_behind_abs']} sayfa geride" if target_progress["pages_ahead_behind"] < 0
                 else "Tam hedefte"),
            ],
        ]
        if target_progress["target_date_passed"]:
            summary_rows.append(["Uyarı", f"Hedef tarih geçti, {target_progress['remaining_pages']} sayfa kaldı"])
        elif target_progress["required_weekly_pace"] is not None:
            summary_rows.append(["Hedefte Kalmak İçin Gereken Tempo", f"Haftada {target_progress['required_weekly_pace']} sayfa"])

    summary_table = Table([summary_headers] + summary_rows, colWidths=[8 * cm, 8 * cm])
    summary_table.setStyle(_table_style([("FONTSIZE", (0, 0), (-1, -1), 9)]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.8 * cm))

    elements.append(Paragraph("Son Ders Kayıtları", styles["Heading3"]))
    lessons = LessonRecord.objects.filter(student=student).order_by("-date")[:20]
    lesson_headers, lesson_rows = lesson_rows_and_headers(lessons)
    lesson_table = Table([lesson_headers] + (lesson_rows or [["Kayıt yok"] + [""] * (len(lesson_headers) - 1)]), repeatRows=1)
    lesson_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#065f46")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
        ("FONTNAME", (0, 1), (-1, -1), FONT_REGULAR),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(lesson_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer
